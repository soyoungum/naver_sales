"""스마트스토어 판매분석·마케팅분석 로데이터 자동 다운로드"""

import argparse
import json
import os
import re
import socket
import subprocess
import sys
import time
from datetime import datetime

# Windows cp949 환경에서 한글/특수문자 출력 깨짐 방지
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf-8-sig'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from pathlib import Path

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# ── 경로 ──────────────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).parent
CRED_FILE    = BASE_DIR / '.smartstore_credentials.json'
DOWNLOADS    = BASE_DIR / 'downloads'


# ══════════════════════════════════════════════════════════════════════════════
# 인증
# ══════════════════════════════════════════════════════════════════════════════
def load_credentials() -> tuple[str, str]:
    try:
        d = json.loads(CRED_FILE.read_text(encoding='utf-8'))
        return str(d.get('login_id', '')), str(d.get('login_pw', ''))
    except Exception:
        return '', ''


def save_credentials(login_id: str, login_pw: str) -> None:
    CRED_FILE.write_text(
        json.dumps({'login_id': login_id, 'login_pw': login_pw},
                   ensure_ascii=False, indent=2),
        encoding='utf-8',
    )


# ══════════════════════════════════════════════════════════════════════════════
# 날짜 파싱
# ══════════════════════════════════════════════════════════════════════════════
def parse_date(raw: str) -> tuple[int, int, int]:
    """YYMMDD 또는 MMDD → (year, month, day)"""
    d = ''.join(c for c in raw if c.isdigit())
    now = datetime.now()
    if len(d) == 6:
        return 2000 + int(d[0:2]), int(d[2:4]), int(d[4:6])
    if len(d) == 4:
        return now.year, int(d[0:2]), int(d[2:4])
    raise ValueError(f'날짜 형식 오류: "{raw}"  →  YYMMDD 또는 MMDD 로 입력하세요.')


# ══════════════════════════════════════════════════════════════════════════════
# Excel 가공
# ══════════════════════════════════════════════════════════════════════════════
ACCESSORY_KEYWORDS = ('커버', '옷걸이', '레버', '받침대', '등판', '스프레이')

# '1세대' 접미사 제외 시리즈 (1세대 표기 무시하고 본 시리즈명만 사용)
NO_GEN1_SUFFIX_SERIES = ('링고',)

# 상품ID별 시리즈명 강제 지정 (악세사리/일반 추출 모두 무시하고 우선 적용)
PRODUCT_ID_OVERRIDE: dict[str, str] = {
    '11609082711': '링고',
    '11838839717': 'T50 AIR',
    '6290832927':  'T50 AIR',
    '12574319352': '악세사리',   # 시디즈 멀티 섬유탈취 스프레이
    '12574313860': '악세사리',   # 시디즈 더 ... (소량 단일 상품)
}


def extract_series(product_name: str) -> str:
    """상품명 → 시리즈명.

    - 악세사리 키워드(커버/옷걸이/레버/받침대/등판/스프레이) 포함 → '악세사리'
    - '1세대' 포함 → 시리즈명 뒤에 ' 1세대' 부착 (예: 'T20' → 'T20 1세대')
        단, NO_GEN1_SUFFIX_SERIES(예: 링고)는 접미사 제외
    - 그 외 → '시디즈 ' 다음 첫 토큰 (예: '시디즈 T50 블랙' → 'T50')
    """
    text = str(product_name or '').strip()
    if any(k in text for k in ACCESSORY_KEYWORDS):
        return '악세사리'
    m = re.search(r'시디즈\s+([^\s,\[]+)', text)
    series = m.group(1).strip() if m else text

    # 누적된 ' 1세대' 접미사 모두 떼고 base만 추출
    while series.endswith(' 1세대'):
        series = series[: -len(' 1세대')].strip()

    if series in NO_GEN1_SUFFIX_SERIES:
        return series

    if '1세대' in text:
        series = f'{series} 1세대'
    return series


def _num(v) -> float:
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(',', ''))
    except ValueError:
        return 0.0


def _save_workbook(wb, path: Path) -> None:
    """OneDrive 동기화 락에 대비한 재시도 + 원자적 저장."""
    tmp = path.with_suffix(path.suffix + '.tmp')
    last_err: Exception | None = None
    for attempt in range(6):
        try:
            wb.save(str(tmp))
            os.replace(tmp, path)
            return
        except OSError as e:
            last_err = e
            wait = 0.5 * (2 ** attempt)   # 0.5, 1, 2, 4, 8, 16초
            print(f'  저장 재시도 {attempt + 1}/6 ({wait:.1f}s 대기) — {e.__class__.__name__}: {e}')
            time.sleep(wait)
    if tmp.exists():
        try:
            tmp.unlink()
        except OSError:
            pass
    raise RuntimeError(
        f'파일 저장 실패: {path.name} — OneDrive 동기화를 일시 중지하고 다시 실행해 주세요.'
    ) from last_err


def process_excel(path: Path) -> None:
    """E열 시리즈명 변환 / U열 실제수량 / V열 실제금액"""
    wb = load_workbook(str(path))
    # 데이터 시트 찾기 (피벗 시트가 아닌 것)
    ws = next((wb[n] for n in wb.sheetnames if n != '피벗_시리즈'), wb.active)

    ws.cell(1, 21).value = '실제수량'
    ws.cell(1, 22).value = '실제금액'

    series_count = 0
    for row in range(2, ws.max_row + 1):
        # E열: 시리즈명 (F열 상품ID 기준 강제 매핑 우선)
        cell = ws.cell(row=row, column=5)
        pid = ws.cell(row=row, column=6).value
        pid_key = str(pid).strip().removesuffix('.0') if pid is not None else ''
        if pid_key in PRODUCT_ID_OVERRIDE:
            new_val = PRODUCT_ID_OVERRIDE[pid_key]
            if cell.value is None or new_val != str(cell.value).strip():
                cell.value = new_val
                series_count += 1
        elif cell.value:
            new_val = extract_series(str(cell.value))
            if new_val != str(cell.value).strip():
                cell.value = new_val
                series_count += 1

        # U열: 실제수량 = H - S
        qty = _num(ws.cell(row=row, column=8).value) - _num(ws.cell(row=row, column=19).value)
        ws.cell(row=row, column=21).value = int(qty) if qty == int(qty) else qty

        # V열: 실제금액 = J - Q
        amt = _num(ws.cell(row=row, column=10).value) - _num(ws.cell(row=row, column=17).value)
        ws.cell(row=row, column=22).value = int(amt) if amt == int(amt) else amt

    _save_workbook(wb, path)
    wb.close()
    print(f'  시리즈명 변환: {series_count}행 / 수량·금액 계산 완료')


def create_pivot(path: Path) -> None:
    """피벗_시리즈 시트: 시리즈별 실제수량·실제금액 합계"""
    wb = load_workbook(str(path))

    # 데이터 시트 찾기 (피벗 시트가 아닌 것)
    data_ws = next((wb[n] for n in wb.sheetnames if n != '피벗_시리즈'), wb.active)

    # 집계
    totals: dict[str, list[float]] = {}  # {시리즈: [수량합, 금액합]}
    for row in range(2, data_ws.max_row + 1):
        series = str(data_ws.cell(row, 5).value or '').strip()
        if not series:
            continue
        qty = _num(data_ws.cell(row, 21).value)
        amt = _num(data_ws.cell(row, 22).value)
        if series not in totals:
            totals[series] = [0.0, 0.0]
        totals[series][0] += qty
        totals[series][1] += amt

    if not totals:
        print('  피벗: 집계할 데이터 없음')
        wb.close()
        return

    # 피벗 시트 초기화
    pname = '피벗_시리즈'
    if pname in wb.sheetnames:
        del wb[pname]
    pws = wb.create_sheet(pname, 0)   # 맨 앞에 추가

    # 헤더
    pws['A1'] = '시리즈'
    pws['B1'] = '실제수량'
    pws['C1'] = '실제금액'

    # 실제금액 내림차순 정렬
    rows = sorted(totals.items(), key=lambda x: x[1][1], reverse=True)
    for i, (series, (qty, amt)) in enumerate(rows, start=2):
        pws.cell(i, 1).value = series
        pws.cell(i, 2).value = round(qty)
        pws.cell(i, 3).value = round(amt)

    _save_workbook(wb, path)
    wb.close()
    print(f'  피벗 생성 완료: {len(rows)}개 시리즈 → 시트: {pname}')


# ══════════════════════════════════════════════════════════════════════════════
# 마케팅분석 Excel 가공
# ══════════════════════════════════════════════════════════════════════════════
def create_marketing_pivot(path: Path) -> None:
    """피벗_채널 시트: 채널명별 고객수·유입수·광고비·결제금액 합계"""
    wb = load_workbook(str(path))

    # 데이터 시트 찾기 (피벗 시트가 아닌 것)
    data_ws = next((wb[n] for n in wb.sheetnames if n != '피벗_채널'), wb.active)

    # 집계: {채널명: [고객수합, 유입수합, 광고비합, 결제금액합]}
    totals: dict[str, list[float]] = {}
    for row in range(2, data_ws.max_row + 1):
        channel = str(data_ws.cell(row, 3).value or '').strip()   # C열: 채널명
        if not channel:
            continue
        customers = _num(data_ws.cell(row, 5).value)              # E열: 고객수
        visits = _num(data_ws.cell(row, 6).value)                 # F열: 유입수
        ad_cost = _num(data_ws.cell(row, 7).value)                # G열: 광고비
        pay_amount = _num(data_ws.cell(row, 12).value)            # L열: 결제금액(마지막클릭)
        if channel not in totals:
            totals[channel] = [0.0, 0.0, 0.0, 0.0]
        totals[channel][0] += customers
        totals[channel][1] += visits
        totals[channel][2] += ad_cost
        totals[channel][3] += pay_amount

    if not totals:
        print('  마케팅 피벗: 집계할 데이터 없음')
        wb.close()
        return

    # 피벗 시트 초기화
    pname = '피벗_채널'
    if pname in wb.sheetnames:
        del wb[pname]
    pws = wb.create_sheet(pname, 0)

    pws['A1'] = '채널명'
    pws['B1'] = '고객수'
    pws['C1'] = '유입수'
    pws['D1'] = '광고비'
    pws['E1'] = '결제금액(마지막클릭)'

    # 유입수 내림차순 정렬
    rows = sorted(totals.items(), key=lambda x: x[1][1], reverse=True)
    for i, (channel, (customers, visits, ad_cost, pay_amount)) in enumerate(rows, start=2):
        pws.cell(i, 1).value = channel
        pws.cell(i, 2).value = round(customers)
        pws.cell(i, 3).value = round(visits)
        pws.cell(i, 4).value = round(ad_cost)
        pws.cell(i, 5).value = round(pay_amount)

    _save_workbook(wb, path)
    wb.close()
    print(f'  마케팅 피벗 생성 완료: {len(rows)}개 채널 → 시트: {pname}')


# ══════════════════════════════════════════════════════════════════════════════
# 캘린더 자동 선택
# ══════════════════════════════════════════════════════════════════════════════
_JS_GET_VISIBLE_MONTHS = """
() => {
    /* 캘린더에 현재 보이는 월 헤더(YYYY.M.) 들 → [[year, month], ...] */
    const isVisible = el => {
        const r = el.getBoundingClientRect();
        if (!r || r.width === 0 || r.height === 0) return false;
        const s = window.getComputedStyle(el);
        return s.visibility !== 'hidden' && s.display !== 'none' && s.opacity !== '0';
    };
    const out = [];
    const seen = new Set();
    document.querySelectorAll('*').forEach(el => {
        if (el.children.length > 0) return;
        const t = (el.textContent || '').replace(/\\s/g, '');
        const m = t.match(/^(\\d{4})\\.(\\d{1,2})\\.?$/);
        if (!m || !isVisible(el)) return;
        const key = m[1] + '.' + m[2];
        if (seen.has(key)) return;
        seen.add(key);
        out.push([parseInt(m[1]), parseInt(m[2])]);
    });
    return out;
}
"""


_JS_CLICK_DAY_IN_PANEL = """
({ year, month, day }) => {
    /* (year, month) 헤더가 있는 패널 안의 day 셀을 클릭.
       캘린더는 한 번에 3개월 패널을 보여주므로, 헤더 X좌표로 패널 경계를 식별. */
    const isVisible = el => {
        const r = el.getBoundingClientRect();
        if (!r || r.width === 0 || r.height === 0) return false;
        const s = window.getComputedStyle(el);
        return s.visibility !== 'hidden' && s.display !== 'none' && s.opacity !== '0';
    };

    // 모든 가시 월 헤더 ('2026. 04.' 또는 '2026.04.' 형태) 수집
    const headers = [];
    document.querySelectorAll('*').forEach(el => {
        if (el.children.length > 0) return;
        const t = (el.textContent || '').replace(/\\s/g, '');
        const m = t.match(/^(\\d{4})\\.(\\d{1,2})\\.?$/);
        if (m && isVisible(el)) {
            headers.push({
                el, year: parseInt(m[1]), month: parseInt(m[2]),
                rect: el.getBoundingClientRect()
            });
        }
    });
    if (!headers.length) return 'NO_HEADERS';

    // 타겟 월 헤더
    const targetH = headers.find(h => h.year === year && h.month === month);
    if (!targetH) {
        const visible = headers.map(h => `${h.year}.${h.month}`).join(',');
        return `NO_HEADER_${year}_${month}_visible:${visible}`;
    }

    // 패널 X 경계: 인접 헤더 사이 중앙
    headers.sort((a, b) => a.rect.left - b.rect.left);
    const idx = headers.indexOf(targetH);
    const minX = idx > 0
        ? (headers[idx-1].rect.right + targetH.rect.left) / 2
        : -Infinity;
    const maxX = idx < headers.length - 1
        ? (targetH.rect.right + headers[idx+1].rect.left) / 2
        : Infinity;

    // 그 패널 안의 day 셀 (텍스트 정확 매치 + Y가 헤더 아래 + X가 panel 범위)
    const dayStr = String(day);
    const candidates = Array.from(document.querySelectorAll('*'))
        .filter(el => el.children.length === 0)
        .filter(el => el.textContent.trim() === dayStr)
        .filter(isVisible)
        .filter(el => {
            const r = el.getBoundingClientRect();
            const cx = r.left + r.width / 2;
            return cx >= minX && cx <= maxX && r.top >= targetH.rect.top;
        });

    if (!candidates.length) {
        return `NO_TARGET_${year}.${month}.${day}_panel:${minX.toFixed(0)}-${maxX.toFixed(0)}`;
    }

    // day <= 7: 다음달 trailing 셀(패널 최하단)과 충돌 가능 → 오름차순(위쪽) 선택
    // day >= 23: 이전달 leading 셀(패널 최상단)과 충돌 가능 → 내림차순(아래쪽) 선택
    // day 8-22: 충돌 없음, 내림차순으로 통일
    if (day <= 7) {
        candidates.sort((a, b) => a.getBoundingClientRect().top - b.getBoundingClientRect().top);
    } else {
        candidates.sort((a, b) => b.getBoundingClientRect().top - a.getBoundingClientRect().top);
    }
    candidates[0].click();
    return `OK_${year}.${month}.${day}_${candidates.length}cells`;
}
"""


def _get_real_frame(page):
    """iframe 요소가 안정된 뒤 실제 Frame 객체 반환"""
    iframe_el = page.wait_for_selector('#__delegate', timeout=15000)
    return iframe_el.content_frame()


_JS_INSPECT_STATE = """
() => {
    const clean = el => (el?.textContent || '').replace(/\\s+/g, ' ').trim();
    // 1) 적용된 날짜 표시
    const toggle = document.querySelector('[data-test-id="DateRangeFixedArea_click_toggle"]');
    const dateText = toggle ? clean(toggle) : null;

    // 2) 다운로드 링크 정보 (모든 후보)
    const links = [];
    document.querySelectorAll('a, button, [role="link"]').forEach(el => {
        if (clean(el) !== '다운로드') return;
        const attrs = {};
        for (const a of el.attributes) attrs[a.name] = a.value;
        links.push({
            tag: el.tagName.toLowerCase(),
            href: el.href || el.getAttribute('href') || null,
            attrs: attrs,
        });
    });

    return { dateText, links };
}
"""


def _inspect_page_state(page) -> dict:
    """페이지에 표시된 날짜 + 다운로드 링크 정보를 캡처."""
    try:
        real_frame = _get_real_frame(page)
        return real_frame.evaluate(_JS_INSPECT_STATE) or {}
    except Exception as e:
        return {'error': str(e)}


def _dates_match(displayed: str | None, sy: int, sm: int, sd: int,
                 ey: int, em: int, ed: int) -> bool:
    """페이지 표시 텍스트에 시작·종료 날짜가 모두 포함되어 있는지 확인.

    스마트스토어는 '2026.04.06.' 또는 '2026.4.6' 형식으로 표시.
    영점 패딩 있는/없는 두 형식 모두 처리.
    """
    if not displayed:
        return False
    # 연.월.일 구분자 제거 후 숫자 스트림
    norm = re.sub(r'[\s.\-~]', '', displayed)
    # 패딩 있는/없는 두 형태 모두 생성
    s_pad   = f'{sy:04d}{sm:02d}{sd:02d}'   # '20260526'
    s_nopad = f'{sy:04d}{sm}{sd}'            # '20260526' or '2026526'
    e_pad   = f'{ey:04d}{em:02d}{ed:02d}'
    e_nopad = f'{ey:04d}{em}{ed}'
    s_ok = (s_pad in norm) or (s_nopad in norm)
    e_ok = (e_pad in norm) or (e_nopad in norm)
    return s_ok and e_ok


def wait_and_verify_dates(page, sy: int, sm: int, sd: int,
                          ey: int, em: int, ed: int,
                          *, max_wait: float = 8.0) -> bool:
    """캘린더 적용 후 페이지가 안정되고 날짜가 적용됐는지 검증.

    - networkidle 대기 → 표시된 날짜 vs 기대 날짜 비교
    - 적용 안 됐으면 max_wait초까지 폴링
    - True 반환이면 검증 통과 / False면 실패 (호출측이 재시도 결정)
    """
    try:
        page.wait_for_load_state('networkidle', timeout=int(max_wait * 1000))
    except Exception as e:
        print(f'  네트워크 idle 대기 시간초과 (계속 진행): {e}')

    deadline = time.time() + max_wait
    while time.time() < deadline:
        state = _inspect_page_state(page)
        displayed = state.get('dateText')
        if _dates_match(displayed, sy, sm, sd, ey, em, ed):
            print(f'  ✓ 날짜 적용 확인: "{displayed}"')
            return True
        time.sleep(0.5)

    state = _inspect_page_state(page)
    print(f'  ✗ 날짜 적용 검증 실패. 표시된 날짜: "{state.get("dateText")}"')
    return False


_JS_EXTRACT_TOTALS = """
() => {
    const cleanText = el => (el?.textContent || '').replace(/\\s+/g, ' ').trim();
    window.scrollTo(0, document.body.scrollHeight);

    const containers = Array.from(
        document.querySelectorAll('table, [role="table"], [role="grid"]'));

    // 1차: 날짜 차원 테이블 — 첫 번째 헤더가 '날짜'인 테이블의 '전체' 행
    //      (날짜별 집계는 중복 없는 dedup 합계 → 진실의 원천)
    const dated = [];
    for (const cont of containers) {
        const headers = Array.from(
            cont.querySelectorAll('th, [role="columnheader"]')).map(cleanText);
        if (headers.length === 0) continue;
        // 첫 헤더가 '날짜' 또는 '날짜↓'/'날짜↑' 형태인 테이블
        if (!headers[0].startsWith('날짜')) continue;
        const rows = Array.from(cont.querySelectorAll('tr, [role="row"]'));
        for (const row of rows) {
            const cells = row.querySelectorAll('td, [role="cell"], [role="gridcell"]');
            if (cells.length < 5) continue;
            if (cleanText(cells[0]) !== '전체') continue;
            dated.push(Array.from(cells).map(cleanText));
        }
    }

    // 2차: '모바일비중' + '상품카테고리' 헤더 테이블의 '전체' 행
    //      (상품카테고리차원 — cross-listing으로 부풀려질 수 있음, 최솟값 선택 시 후순위)
    const anchored = [];
    for (const cont of containers) {
        const headers = Array.from(
            cont.querySelectorAll('th, [role="columnheader"]')).map(cleanText);
        const headerJoin = headers.join(' | ');
        if (!headerJoin.includes('모바일비중')) continue;
        if (!headerJoin.includes('상품카테고리')) continue;
        const rows = Array.from(cont.querySelectorAll('tr, [role="row"]'));
        for (const row of rows) {
            const cells = row.querySelectorAll('td, [role="cell"], [role="gridcell"]');
            if (cells.length < 5) continue;
            if (cleanText(cells[0]) !== '전체') continue;
            anchored.push(Array.from(cells).map(cleanText));
        }
    }

    // 3차: 패턴 매칭 fallback (선두 '전체' N개 + 데이터 5~9개 셀)
    const pattern = [];
    const allRows = Array.from(document.querySelectorAll('tr, [role="row"]'));
    for (const row of allRows) {
        const cells = row.querySelectorAll('td, [role="cell"], [role="gridcell"]');
        if (cells.length < 5 || cells.length > 12) continue;
        const texts = Array.from(cells).map(cleanText);
        let lead = 0;
        for (const t of texts) {
            if (t === '전체') lead++; else break;
        }
        if (lead === 0) continue;
        const data = cells.length - lead;
        if (data < 5 || data > 9) continue;
        pattern.push(texts);
    }

    return { dated, anchored, pattern };
}
"""


def _parse_totals_row(cells: list[str]) -> dict | None:
    """행 셀 리스트에서 (결제금액, 환불금액) 추출. 검증 실패 시 None."""
    nums = []
    for s in cells:
        if not s or s == '전체' or '%' in s:
            continue
        digits = re.sub(r'[^0-9]', '', s)
        if digits:
            nums.append(int(digits))
    if len(nums) < 2:
        return None
    nums_sorted = sorted(nums, reverse=True)
    pay, ref = nums_sorted[0], nums_sorted[1]
    if ref >= pay or pay < 100_000_000:
        return None
    return {'결제금액': pay, '환불금액': ref}


def extract_product_totals(page) -> dict | None:
    """상품성과 페이지 '전체' 합계 행에서 결제금액·환불금액 추출.

    엑셀 다운로드는 카테고리 cross-listing으로 행이 중복 부풀려져 있어
    페이지 값과 다름. 페이지 날짜차원 합계가 진실의 원천이므로 별도 캡처.

    전략: dated(날짜차원) → anchored(상품카테고리차원) → pattern(패턴매칭) 순으로
    수집한 뒤, 모든 후보 중 최소 결제금액을 채택.
    — dedup 된 진짜 합계는 cross-listed 값보다 항상 작음.
    """
    try:
        real_frame = _get_real_frame(page)
        result = real_frame.evaluate(_JS_EXTRACT_TOTALS) or {}
    except Exception as e:
        print(f'  푸터 합계 추출 실패 (frame): {e}')
        return None

    dated = result.get('dated', []) if isinstance(result, dict) else []
    anchored = result.get('anchored', []) if isinstance(result, dict) else []
    pattern = result.get('pattern', []) if isinstance(result, dict) else []

    print(f'  스크래핑 결과: dated={len(dated)}건 / anchored={len(anchored)}건 / pattern={len(pattern)}건')

    # 모든 전략의 후보를 합산 후 최솟값 채택
    # — dedup 된 진짜 합계는 cross-listed 값보다 항상 작음
    all_parsed: list[tuple[dict, list[str], str]] = []
    for cells in dated:
        p = _parse_totals_row(cells)
        if p:
            all_parsed.append((p, cells, 'dated'))
        else:
            print(f'    dated 검증 실패: cells={cells}')
    for cells in anchored:
        p = _parse_totals_row(cells)
        if p:
            all_parsed.append((p, cells, 'anchored'))
        else:
            print(f'    anchored 검증 실패: cells={cells}')
    for cells in pattern:
        p = _parse_totals_row(cells)
        if p:
            all_parsed.append((p, cells, 'pattern'))

    if not all_parsed:
        print('  ✗ 푸터 합계 추출 실패: 모든 후보 검증 실패')
        for label, cands in [('dated', dated), ('anchored', anchored), ('pattern', pattern)]:
            for cells in cands:
                print(f'    [{label}] raw cells: {cells}')
        return None

    all_parsed.sort(key=lambda x: x[0]['결제금액'])
    chosen, _, src = all_parsed[0]
    print(f'  ✓ 채택 ({src}, 전체 {len(all_parsed)}건 중 최솟값): '
          f'결제 {chosen["결제금액"]:,} / 환불 {chosen["환불금액"]:,}')
    if len(all_parsed) > 1:
        for i, (p, cells, s) in enumerate(all_parsed):
            mark = ' ← 선택' if i == 0 else ''
            print(f'    [{i}] {s}: 결제 {p["결제금액"]:,} / 환불 {p["환불금액"]:,}{mark}')

    return chosen


def select_date_range(page, frame, sy: int, sm: int, sd: int,
                      ey: int, em: int, ed: int) -> None:
    """캘린더에서 날짜 범위 선택 후 적용.

    page  : Page          – 최신 Frame 재취득용
    frame : FrameLocator  – click / wait_for 용
    """

    # 캘린더 열기
    print('  캘린더 열기...')
    frame.locator('[data-test-id="DateRangeFixedArea_click_toggle"]').click(timeout=10000)
    frame.locator('[data-test-id="DateCommonPickInfo_click_complate"]').wait_for(
        state='visible', timeout=10000
    )

    # 캘린더 열린 뒤 Frame 재취득 (열기 전 Frame은 detach될 수 있음)
    real_frame = _get_real_frame(page)

    # 이전 달 이동 — 시작·종료 월 둘 다 보일 때까지 동적 클릭
    # (판매분석/마케팅분석 default 상태가 달라서 고정 횟수 못 씀)
    btn = frame.locator('[data-test-id="DateCommonNavBarM_click_prev-month"]').first
    nav_clicks = 0
    for _ in range(13):
        visible = real_frame.evaluate(_JS_GET_VISIBLE_MONTHS) or []
        visible_set = {(m[0], m[1]) for m in visible}
        if (sy, sm) in visible_set and (ey, em) in visible_set:
            break
        btn.click(timeout=3000)
        time.sleep(0.4)
        real_frame = _get_real_frame(page)
        nav_clicks += 1
    else:
        raise RuntimeError(
            f'캘린더 navigation 실패 — {sy}-{sm} / {ey}-{em} 가시되지 않음. '
            f'마지막 visible: {visible}'
        )
    print(f'  이전달 버튼 {nav_clicks}회 클릭 (visible: {visible})')

    # 시작 월 헤더 클릭으로 그 월 패널 활성화 (사용자 검증 패턴)
    header_text = f'{sy}. {sm:02d}.'
    print(f'  시작 월 헤더 클릭: "{header_text}"')
    try:
        frame.get_by_text(header_text, exact=True).first.click(timeout=3000)
        time.sleep(0.3)
    except Exception as e:
        print(f'    헤더 클릭 실패 (계속 진행): {e}')

    # 시작일 클릭 — JS로 정확한 패널 식별
    print(f'  시작일 {sy}-{sm:02d}-{sd:02d}')
    result = real_frame.evaluate(_JS_CLICK_DAY_IN_PANEL,
                                 {'year': sy, 'month': sm, 'day': sd})
    print(f'    {result}')
    if not result.startswith('OK'):
        raise RuntimeError(f'시작일({sy}-{sm}-{sd}) 클릭 실패: {result}')
    time.sleep(0.4)

    # 종료일 클릭 — 다른 월일 수도 있음 (예: 2/23 ~ 3/1)
    print(f'  종료일 {ey}-{em:02d}-{ed:02d}')
    result = real_frame.evaluate(_JS_CLICK_DAY_IN_PANEL,
                                 {'year': ey, 'month': em, 'day': ed})
    print(f'    {result}')
    if not result.startswith('OK'):
        raise RuntimeError(f'종료일({ey}-{em}-{ed}) 클릭 실패: {result}')
    time.sleep(0.4)

    # 적용
    print('  적용 버튼 클릭')
    frame.locator('[data-test-id="DateCommonPickInfo_click_complate"]').click(timeout=5000)
    time.sleep(1.0)


# ══════════════════════════════════════════════════════════════════════════════
# 대시보드 실행
# ══════════════════════════════════════════════════════════════════════════════
def launch_dashboard(port: int = 8501) -> None:
    """Streamlit 서버가 안 떠 있으면 띄우고, Chrome으로 대시보드 열기."""
    url = f'http://localhost:{port}'
    dashboard_script = BASE_DIR / 'dashboard_app.py'

    def port_alive() -> bool:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(0.3)
            return s.connect_ex(('127.0.0.1', port)) == 0

    if not port_alive():
        if not dashboard_script.exists():
            print(f'  대시보드 스크립트 없음: {dashboard_script}')
            return
        print('  Streamlit 서버 시작 중...')
        flags = 0
        if os.name == 'nt':
            flags = (subprocess.CREATE_NEW_PROCESS_GROUP
                     | getattr(subprocess, 'DETACHED_PROCESS', 0))
        subprocess.Popen(
            [sys.executable, '-m', 'streamlit', 'run', str(dashboard_script),
             '--server.port', str(port),
             '--server.headless', 'true',
             '--browser.gatherUsageStats', 'false'],
            cwd=str(BASE_DIR),
            creationflags=flags,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, stdin=subprocess.DEVNULL,
            close_fds=True,
        )
        # 서버 준비 대기 (최대 15초)
        ready = False
        for _ in range(30):
            time.sleep(0.5)
            if port_alive():
                ready = True
                break
        if not ready:
            print('  서버 시작 실패 — 수동으로 실행해 주세요')
            return

    print(f'  Chrome으로 열기: {url}')
    try:
        subprocess.Popen(['cmd', '/c', 'start', '', 'chrome', url])
    except FileNotFoundError:
        import webbrowser
        webbrowser.open(url)


# ══════════════════════════════════════════════════════════════════════════════
# 메인
# ══════════════════════════════════════════════════════════════════════════════
def main() -> None:
    parser = argparse.ArgumentParser(description='스마트스토어 상품성과 다운로드')
    parser.add_argument('--start',    default='', help='시작일 (YYMMDD 또는 MMDD)')
    parser.add_argument('--end',      default='', help='종료일 (YYMMDD 또는 MMDD)')
    parser.add_argument('--target',   type=float, default=None,
                        help='목표 거래액 (억 단위, 예: 9)')
    parser.add_argument('--headless', action='store_true')
    parser.add_argument('--force',    action='store_true',
                        help='이미 받은 주차도 강제로 다시 받기 (기본은 기존 파일 보존)')
    args = parser.parse_args()

    # ── 인증 정보 ──
    saved_id, saved_pw = load_credentials()
    login_id = os.getenv('SMARTSTORE_ID', '').strip() or saved_id
    login_pw = os.getenv('SMARTSTORE_PW', '').strip() or saved_pw
    if not login_id:
        login_id = input('아이디: ').strip()
    if not login_pw:
        login_pw = input('비밀번호: ').strip()
    if not saved_id or not saved_pw:
        save_credentials(login_id, login_pw)

    # ── 날짜 / 목표 입력 ──
    if not args.start or not args.end:
        print('\n[네이버 행사 주차 입력]')
    start_raw = args.start or input('시작일 (YYMMDD): ').strip()
    end_raw   = args.end   or input('종료일 (YYMMDD): ').strip()
    sy, sm, sd = parse_date(start_raw)
    ey, em, ed = parse_date(end_raw)

    # 목표 거래액 (억 단위)
    if args.target is not None:
        target_eok = args.target
    elif args.start and args.end:
        target_eok = None
    else:
        t_str = input('목표 거래액 (억, 빈 값이면 스킵): ').strip()
        try:
            target_eok = float(t_str) if t_str else None
        except ValueError:
            print(f'  목표 입력 오류 — "{t_str}" 무시하고 진행')
            target_eok = None

    DOWNLOADS.mkdir(exist_ok=True)
    sales_path     = DOWNLOADS / f'sales-{start_raw}-{end_raw}.xlsx'
    marketing_path = DOWNLOADS / f'marketing-{start_raw}-{end_raw}.xlsx'

    # 목표 즉시 저장 (다운로드 실패해도 보존)
    if target_eok is not None:
        target_path = DOWNLOADS / f'sales-{start_raw}-{end_raw}.target.json'
        target_path.write_text(
            json.dumps({'target_eok': target_eok}, ensure_ascii=False, indent=2),
            encoding='utf-8',
        )
        print(f'  목표 거래액 저장: {target_eok}억 → {target_path.name}')

    # ── 기존 파일 보호 (소급 반영 방지) ──
    if not args.force and sales_path.exists() and marketing_path.exists():
        print(f'\n해당 주차 파일이 이미 존재합니다 — 다운로드 건너뜁니다 (소급 반영 방지):')
        print(f'  {sales_path.name}')
        print(f'  {marketing_path.name}')
        print(f'다시 받으려면: --force 옵션 사용 또는 기존 파일을 직접 삭제 후 재실행')
        print('\n대시보드 여는 중...')
        launch_dashboard()
        return

    # ── 브라우저 ──
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=args.headless)
        ctx     = browser.new_context(accept_downloads=True)
        page    = ctx.new_page()

        # 로그인
        print('로그인 중...')
        page.goto('https://sell.smartstore.naver.com/home')
        page.wait_for_load_state('networkidle')
        page.locator('button:has-text("로그인하기"), a:has-text("로그인하기")').first.click()
        page.get_by_role('textbox', name='아이디 또는 이메일 주소').fill(login_id)
        page.get_by_role('textbox', name='비밀번호').fill(login_pw)
        page.get_by_role('button', name='로그인', exact=True).click()

        # 캡차 (발생 시)
        try:
            captcha = page.get_by_role('textbox', name='자동입력 방지 문자')
            if captcha.is_visible(timeout=6000):
                code = input('캡차 입력: ').strip()
                captcha.fill(code)
                page.get_by_role('button', name='로그인', exact=True).click()
        except Exception:
            pass

        # ── 1) 판매분석 → 상품성과 다운로드 ──
        print('\n[1/2] 데이터분석 → 판매분석 → 상품성과 이동 중...')
        page.get_by_role('menuitem', name='데이터분석').click()
        page.get_by_role('link', name='판매분석').click()
        page.get_by_role('link', name='상품성과').click()

        frame = page.frame_locator('#__delegate')

        print(f'날짜 선택: {sy}-{sm:02d}-{sd:02d} ~ {ey}-{em:02d}-{ed:02d}')
        select_date_range(page, frame, sy, sm, sd, ey, em, ed)

        # 적용 검증 + 자동 재시도 (최대 1회)
        if not wait_and_verify_dates(page, sy, sm, sd, ey, em, ed):
            print('  → 캘린더 재선택 후 1회 재시도')
            select_date_range(page, frame, sy, sm, sd, ey, em, ed)
            if not wait_and_verify_dates(page, sy, sm, sd, ey, em, ed):
                raise RuntimeError(
                    f'날짜 적용 실패: {sy}-{sm:02d}-{sd:02d} ~ {ey}-{em:02d}-{ed:02d}\n'
                    '잘못된 날짜로 데이터를 받으면 안 되므로 중단합니다.'
                )

        # 다운로드 직전 상태 디버그 로깅
        pre_state = _inspect_page_state(page)
        print(f'  [DEBUG] 다운로드 직전 표시 날짜: "{pre_state.get("dateText")}"')
        for i, lk in enumerate(pre_state.get('links') or []):
            print(f'  [DEBUG] 다운로드 링크[{i}] tag={lk.get("tag")} href={lk.get("href")}')
            print(f'  [DEBUG]                 attrs={lk.get("attrs")}')

        print('판매분석 다운로드 중...')
        with page.expect_download(timeout=30000) as dl_info:
            frame.get_by_role('link', name='다운로드').click()
        dl = dl_info.value
        dl.save_as(str(sales_path))
        print(f'저장: {sales_path}')

        # 페이지 푸터 '전체' 합계 캡처 (엑셀 행 중복 보정용 — 진실의 원천)
        # 다운로드 클릭 후 페이지가 날짜 필터를 리셋할 수 있으므로 날짜를 재확인
        time.sleep(1.0)
        date_ok = wait_and_verify_dates(page, sy, sm, sd, ey, em, ed, max_wait=5.0)
        if not date_ok:
            print('  ⚠ 다운로드 후 날짜 필터가 리셋됨 — 날짜 재선택 후 totals 추출')
            real_frame = _get_real_frame(page)
            select_date_range(page, frame, sy, sm, sd, ey, em, ed)
            wait_and_verify_dates(page, sy, sm, sd, ey, em, ed)
        totals = extract_product_totals(page)
        if totals:
            totals_path = DOWNLOADS / f'sales-{start_raw}-{end_raw}.totals.json'
            totals_path.write_text(
                json.dumps(totals, ensure_ascii=False, indent=2),
                encoding='utf-8',
            )
            print(f'  푸터 합계 저장: 결제 {totals["결제금액"]:,} / 환불 {totals["환불금액"]:,}')

        # ── 2) 마케팅분석 다운로드 ──
        print('\n[2/2] 데이터분석 → 마케팅분석 이동 중...')
        page.get_by_role('menuitem', name='데이터분석').click()
        page.get_by_role('link', name='마케팅분석').click()
        time.sleep(2)

        frame = page.frame_locator('#__delegate')

        print(f'날짜 선택: {sy}-{sm:02d}-{sd:02d} ~ {ey}-{em:02d}-{ed:02d}')
        select_date_range(page, frame, sy, sm, sd, ey, em, ed)

        # 적용 검증 + 자동 재시도 (최대 1회)
        if not wait_and_verify_dates(page, sy, sm, sd, ey, em, ed):
            print('  → 캘린더 재선택 후 1회 재시도')
            select_date_range(page, frame, sy, sm, sd, ey, em, ed)
            if not wait_and_verify_dates(page, sy, sm, sd, ey, em, ed):
                raise RuntimeError(
                    f'날짜 적용 실패: {sy}-{sm:02d}-{sd:02d} ~ {ey}-{em:02d}-{ed:02d}\n'
                    '잘못된 날짜로 데이터를 받으면 안 되므로 중단합니다.'
                )

        pre_state = _inspect_page_state(page)
        print(f'  [DEBUG] 다운로드 직전 표시 날짜: "{pre_state.get("dateText")}"')
        for i, lk in enumerate(pre_state.get('links') or []):
            print(f'  [DEBUG] 다운로드 링크[{i}] tag={lk.get("tag")} href={lk.get("href")}')
            print(f'  [DEBUG]                 attrs={lk.get("attrs")}')

        print('마케팅분석 다운로드 중...')
        with page.expect_download(timeout=30000) as dl_info:
            frame.get_by_role('link', name='다운로드').click()
        dl = dl_info.value
        dl.save_as(str(marketing_path))
        print(f'저장: {marketing_path}')

        ctx.close()
        browser.close()

    # ── Excel 가공 ──
    print('\n판매분석 Excel 처리 중...')
    process_excel(sales_path)
    create_pivot(sales_path)
    print(f'완료 → {sales_path}')

    print('\n마케팅분석 Excel 처리 중...')
    create_marketing_pivot(marketing_path)
    print(f'완료 → {marketing_path}')

    print('\n대시보드 여는 중...')
    launch_dashboard()


if __name__ == '__main__':
    main()
