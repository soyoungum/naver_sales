"""수기 다운로드 vs 자동 다운로드 대조 스크립트

사용법:
  python compare_data.py <수기파일.xlsx> <자동파일.xlsx>

파일을 downloads/ 폴더에 넣고 실행하거나,
인자 없이 실행하면 downloads/ 폴더의 최신 파일 2개를 자동으로 씁니다.
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR  = Path(__file__).parent
DOWNLOADS = BASE_DIR / 'downloads'

PIVOT_SHEET = '피벗_시리즈'
DATA_SHEET_EXCLUDE = PIVOT_SHEET


def load_pivot(path: Path) -> dict[str, dict]:
    """피벗_시리즈 시트에서 {시리즈: {qty, amt}} 로드"""
    wb = load_workbook(str(path))
    if PIVOT_SHEET in wb.sheetnames:
        ws = wb[PIVOT_SHEET]
        data = {}
        for r in range(2, ws.max_row + 1):
            series = str(ws.cell(r, 1).value or '').strip()
            qty = ws.cell(r, 2).value or 0
            amt = ws.cell(r, 3).value or 0
            if series:
                data[series] = {'qty': float(qty), 'amt': float(amt)}
        wb.close()
        return data
    wb.close()
    return {}


def load_raw(path: Path) -> dict[str, dict]:
    """데이터 시트(상품성과)에서 {상품명: {qty, amt}} 로드"""
    wb = load_workbook(str(path))
    ws = next((wb[n] for n in wb.sheetnames if n != DATA_SHEET_EXCLUDE), wb.active)
    data = {}
    for r in range(2, ws.max_row + 1):
        name = str(ws.cell(r, 5).value or '').strip()
        if not name:
            continue
        qty = float(str(ws.cell(r, 8).value or 0).replace(',', '') or 0)
        amt = float(str(ws.cell(r, 10).value or 0).replace(',', '') or 0)
        if name not in data:
            data[name] = {'qty': 0.0, 'amt': 0.0}
        data[name]['qty'] += qty
        data[name]['amt'] += amt
    wb.close()
    return data


def compare(auto_path: Path, manual_path: Path) -> None:
    print(f'\n[자동] {auto_path.name}')
    print(f'[수기] {manual_path.name}')

    # 피벗 비교 (있으면)
    auto_pivot  = load_pivot(auto_path)
    manual_pivot = load_pivot(manual_path)

    if auto_pivot and manual_pivot:
        print('\n=== 피벗(시리즈별) 대조 ===')
        print(f'{"시리즈":<30} {"자동수량":>8} {"수기수량":>8} {"수량차":>6}  {"자동금액":>14} {"수기금액":>14} {"금액차":>12}')
        print('-'*100)
        all_keys = sorted(auto_pivot.keys() | manual_pivot.keys())
        diff_count = 0
        for k in all_keys:
            a = auto_pivot.get(k, {'qty': 0, 'amt': 0})
            m = manual_pivot.get(k, {'qty': 0, 'amt': 0})
            dq = round(a['qty']) - round(m['qty'])
            da = round(a['amt']) - round(m['amt'])
            marker = ' ◀ 불일치' if dq != 0 or da != 0 else ''
            if dq != 0 or da != 0:
                diff_count += 1
            print(f'{k:<30} {round(a["qty"]):>8,} {round(m["qty"]):>8,} {dq:>+6,}  {round(a["amt"]):>14,} {round(m["amt"]):>14,} {da:>+12,}{marker}')
        if diff_count == 0:
            print('\n✓ 모든 시리즈 일치!')
        else:
            print(f'\n✗ 불일치 시리즈: {diff_count}개')
    else:
        # 피벗 없으면 상품명 기준으로 결제금액·수량 비교
        print('\n=== 상품성과(상품별 결제수량·결제금액) 대조 ===')
        auto_raw   = load_raw(auto_path)
        manual_raw = load_raw(manual_path)
        all_keys = sorted(auto_raw.keys() | manual_raw.keys())
        diff_count = 0
        for k in all_keys:
            a = auto_raw.get(k, {'qty': 0, 'amt': 0})
            m = manual_raw.get(k, {'qty': 0, 'amt': 0})
            dq = round(a['qty']) - round(m['qty'])
            da = round(a['amt']) - round(m['amt'])
            if dq != 0 or da != 0:
                diff_count += 1
                print(f'  [{k}] 수량차={dq:+,}  금액차={da:+,}')
        if diff_count == 0:
            print('✓ 모든 상품 일치!')
        else:
            print(f'\n✗ 불일치 상품: {diff_count}개')

    # 전체 합계 비교
    print('\n=== 전체 합계 ===')
    ap = load_pivot(auto_path) or {k: v for d in [load_raw(auto_path)] for k, v in d.items()}
    mp = load_pivot(manual_path) or {k: v for d in [load_raw(manual_path)] for k, v in d.items()}
    a_qty = round(sum(v['qty'] for v in ap.values()))
    a_amt = round(sum(v['amt'] for v in ap.values()))
    m_qty = round(sum(v['qty'] for v in mp.values()))
    m_amt = round(sum(v['amt'] for v in mp.values()))
    print(f'  자동: 수량 {a_qty:,}  금액 {a_amt:,}')
    print(f'  수기: 수량 {m_qty:,}  금액 {m_amt:,}')
    print(f'  차이: 수량 {a_qty - m_qty:+,}  금액 {a_amt - m_amt:+,}')


if __name__ == '__main__':
    if len(sys.argv) == 3:
        auto_path   = Path(sys.argv[1])
        manual_path = Path(sys.argv[2])
    else:
        xlsx_files = sorted(DOWNLOADS.glob('*.xlsx'), key=lambda p: p.stat().st_mtime)
        if len(xlsx_files) < 2:
            print('downloads/ 폴더에 xlsx 파일이 2개 이상 필요합니다.')
            print('사용법: python compare_data.py <수기파일.xlsx> <자동파일.xlsx>')
            sys.exit(1)
        auto_path   = xlsx_files[-1]
        manual_path = xlsx_files[-2]
        print(f'자동 선택: {auto_path.name} vs {manual_path.name}')

    compare(auto_path, manual_path)
